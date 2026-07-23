"""
app/api.py

Api — the sole bridge surface between the React frontend and the Python backend.

Every method on this class is callable from React via:
    window.pywebview.api.method_name(args)

Methods are grouped into three categories:

  PULL (called once on React mount to hydrate state):
    get_initial_state()

  DIALOG (replace native Qt file dialogs):
    pick_pipeline_folder()
    pick_input_file()
    pick_output_file()
    pick_python_exe()

  QUERY:
    get_sheet_names(file_path)
    test_python(exe)

  COMMAND (mirror what used to be Qt slots):
    save_settings(settings)
    start_pipeline(settings)
    pause_pipeline()
    resume_pipeline()
    cancel_pipeline()
    confirm_login()
    open_file(path)
    open_log(pipeline_dir)

  PUSH (called by PipelineRunner from its background thread):
    push_log(payload)
    push_progress(payload)
    push_done(payload)
    push_error(payload)
    push_awaiting_login()

Push payloads are serialised to JSON, base64-encoded, and delivered via
window.evaluate_js() as CustomEvents.  The base64 encoding avoids all JS
string-escaping edge cases.  The JS side decodes with Uint8Array + TextDecoder
(not bare atob()) to handle multi-byte UTF-8 characters in Marathi college
names correctly.
"""
from __future__ import annotations

import base64
import glob
import json
import os
import subprocess
from typing import Optional

import openpyxl
import webview

from app.settings import (
    Settings,
    default_pipeline_dir,
    load_settings,
    save_settings,
)
from app.runner import PipelineRunner


def _count_input_rows(settings_dict: dict) -> Optional[int]:
    """
    Count the number of data rows that will actually be processed.

    In 'all' mode this is the full sheet minus the header row.
    In 'range' mode this is  end_row - start_row + 1  clamped to the
    actual sheet size (both values are 1-indexed, header = row 1).

    Returns None if anything goes wrong — the Run screen handles None
    gracefully by showing an indeterminate progress counter.
    """
    try:
        wb = openpyxl.load_workbook(settings_dict.get("input_file", ""), read_only=True)
        ws = wb[settings_dict.get("sheet_name", "")]
        sheet_data_rows = max((ws.max_row or 1) - 1, 0)
        wb.close()

        if settings_dict.get("process_mode") == "range":
            try:
                start = int(settings_dict.get("start_row", 0))
                end = int(settings_dict.get("end_row", 0))
                range_count = max(end - start + 1, 0)
                return min(range_count, sheet_data_rows) or None
            except (ValueError, TypeError):
                pass

        return sheet_data_rows or None
    except Exception:
        return None


class Api:
    """JS bridge surface exposed as window.pywebview.api in the React frontend."""

    def __init__(self) -> None:
        self._window: Optional[webview.Window] = None
        self._runner: Optional[PipelineRunner] = None

    def set_window(self, window: webview.Window) -> None:
        """Called by main.py immediately after webview.create_window()."""
        self._window = window

    # ------------------------------------------------------------------
    # PULL — called on React mount to hydrate state
    # ------------------------------------------------------------------

    def get_initial_state(self) -> dict:
        """
        Return everything the Settings screen needs on first load.
        Fixes the previously-open bug where Pipeline Folder was not
        auto-populated on first launch after the pipeline installer ran.
        """
        return {
            "settings": load_settings(),
            "default_pipeline_dir": default_pipeline_dir(),
        }

    # ------------------------------------------------------------------
    # DIALOG — replace native Qt file dialogs
    # ------------------------------------------------------------------

    def pick_pipeline_folder(self) -> Optional[str]:
        """Open a folder-picker dialog and return the selected path, or None."""
        result = self._window.create_file_dialog(webview.FOLDER_DIALOG)
        if result:
            return result[0] if isinstance(result, (list, tuple)) else result
        return None

    def pick_input_file(self) -> Optional[str]:
        """Open an open-file dialog filtered to Excel files."""
        result = self._window.create_file_dialog(
            webview.OPEN_DIALOG,
            file_types=("Excel Files (*.xlsx;*.xls)", "All Files (*.*)"),
        )
        if result:
            return result[0] if isinstance(result, (list, tuple)) else result
        return None

    def pick_output_file(self) -> Optional[str]:
        """Open a save-file dialog for the output Excel path."""
        result = self._window.create_file_dialog(
            webview.SAVE_DIALOG,
            file_types=("Excel Files (*.xlsx)", "All Files (*.*)"),
        )
        if result:
            path = result[0] if isinstance(result, (list, tuple)) else result
            if path and not path.endswith(".xlsx"):
                path += ".xlsx"
            return path
        return None

    def pick_python_exe(self) -> Optional[str]:
        """Open an open-file dialog filtered to executables."""
        result = self._window.create_file_dialog(
            webview.OPEN_DIALOG,
            file_types=("Executables (*.exe)", "All Files (*.*)"),
        )
        if result:
            return result[0] if isinstance(result, (list, tuple)) else result
        return None

    # ------------------------------------------------------------------
    # QUERY
    # ------------------------------------------------------------------

    def get_sheet_names(self, file_path: str) -> list:
        """
        Read sheet names from an Excel workbook.
        Returns an empty list if the file is unreadable or not an Excel file.
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception:
            return []

    def test_python(self, exe: str) -> dict:
        """
        Run `exe --version` and return the result.
        Returns {"ok": bool, "version": str, "error": str}.
        """
        exe = exe.strip() or "python"
        try:
            result = subprocess.run(
                [exe, "--version"],
                capture_output=True,
                text=True,
                timeout=10,
            )
            version = result.stdout.strip() or result.stderr.strip()
            return {"ok": True, "version": version, "error": ""}
        except FileNotFoundError:
            return {"ok": False, "version": "", "error": f"Interpreter not found: {exe}"}
        except subprocess.TimeoutExpired:
            return {"ok": False, "version": "", "error": "Timed out"}

    # ------------------------------------------------------------------
    # COMMAND — mirror what used to be Qt slots
    # ------------------------------------------------------------------

    def save_settings(self, settings: dict) -> dict:
        """Persist settings to disk.  Called from Settings.jsx on any save."""
        save_settings(settings)
        return {"ok": True}

    def start_pipeline(self, settings: dict) -> dict:
        """
        Persist settings, count rows for the progress bar, then launch the runner.

        Returns {"ok": True, "total_rows": int | None} so the Run screen
        can initialise a determinate or indeterminate progress bar without a
        separate pull call.
        """
        save_settings(settings)
        total_rows = _count_input_rows(settings)

        # Inject total_rows_hint into the settings dict so PipelineRunner
        # can read it without a separate constructor argument.
        enriched = dict(settings)
        enriched["_total_rows_hint"] = total_rows

        self._runner = PipelineRunner(api_ref=self, settings=enriched)
        self._runner.start()
        return {"ok": True, "total_rows": total_rows}

    def pause_pipeline(self) -> None:
        """Suspend the running subprocess (psutil SuspendThread on Windows)."""
        if self._runner:
            self._runner.pause()

    def resume_pipeline(self) -> None:
        """Resume a previously suspended subprocess."""
        if self._runner:
            self._runner.resume()

    def cancel_pipeline(self) -> None:
        """Terminate the subprocess and all its children (kills Playwright's chrome.exe)."""
        if self._runner:
            self._runner.cancel()

    def confirm_login(self) -> None:
        """
        Unblock the pipeline's input() call after the user has logged in to the
        MAHABOCW portal.  Sends a newline to the subprocess stdin.
        """
        if self._runner:
            self._runner.confirm_login()

    def open_file(self, path: str) -> None:
        """Open a file with its default Windows application (os.startfile)."""
        if path and os.path.isfile(path):
            os.startfile(path)

    def open_log(self, pipeline_dir: str) -> None:
        """Open the most recent .log file from the pipeline's logs/ folder."""
        if not pipeline_dir:
            return
        log_dir = os.path.join(pipeline_dir, "logs")
        log_files = glob.glob(os.path.join(log_dir, "*.log"))
        if log_files:
            latest = max(log_files, key=os.path.getmtime)
            os.startfile(latest)

    # ------------------------------------------------------------------
    # PUSH — called by PipelineRunner from the background thread
    # ------------------------------------------------------------------

    def push_log(self, payload: dict) -> None:
        """Dispatch a 'pipeline-log' event to the React frontend."""
        self._dispatch("pipeline-log", payload)

    def push_progress(self, payload: dict) -> None:
        """Dispatch a 'pipeline-progress' event with count and total."""
        self._dispatch("pipeline-progress", payload)

    def push_done(self, payload: dict) -> None:
        """
        Dispatch a 'pipeline-done' event on successful completion.
        payload: {"claims_seen": int, "output_file": str}
        """
        self._dispatch("pipeline-done", payload)

    def push_error(self, payload: dict) -> None:
        """
        Dispatch a 'pipeline-error' event on failure or cancellation.
        payload: {"message": str}
        """
        self._dispatch("pipeline-error", payload)

    def push_awaiting_login(self) -> None:
        """
        Dispatch a 'pipeline-awaiting-login' event — fired once when the
        pipeline's input() prompt appears.  The Run screen shows the login
        confirmation banner in response.
        """
        self._dispatch("pipeline-awaiting-login", {})

    # ------------------------------------------------------------------
    # Internal dispatch helper
    # ------------------------------------------------------------------

    def _dispatch(self, event_name: str, payload: dict) -> None:
        """
        Serialise *payload* to JSON, base64-encode it, and inject it into the
        browser as a CustomEvent via window.evaluate_js().

        Base64 avoids all JS string-escaping issues (backslashes, quotes,
        newlines in log lines, etc.).

        The JS decodes with Uint8Array + TextDecoder instead of bare atob()
        so multi-byte UTF-8 characters (Marathi/Devanagari college names in
        log lines) are decoded correctly.

        evaluate_js() is safe to call from a non-main thread in pywebview.
        """
        if not self._window:
            return

        json_str = json.dumps(payload, ensure_ascii=False)
        b64_str = base64.b64encode(json_str.encode("utf-8")).decode("ascii")

        js = f"""
        (function() {{
            var b64 = '{b64_str}';
            var bytes = Uint8Array.from(atob(b64), function(c) {{ return c.charCodeAt(0); }});
            var detail = JSON.parse(new TextDecoder('utf-8').decode(bytes));
            window.dispatchEvent(new CustomEvent('{event_name}', {{ detail: detail }}));
        }})();
        """
        self._window.evaluate_js(js)
